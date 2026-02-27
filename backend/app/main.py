import openpyxl
import asyncio
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import re
from calendar import monthrange
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, send_from_directory, session, g
from flask_cors import CORS
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import shutil
from config.settings import Config
from models.physchem import db, PhysChemAnalysis
from models.microbiological import MicrobiologicalAnalysis
from models.water_treatment import WaterTreatmentReading
from models.dam_level_snapshot import DamLevelSnapshot
from models.turbidity_snapshot import TurbiditySnapshot
from models.leave_records import CTOApplication, LeaveApplication, LeaveCredits, Employee
from models.auth import AppUser
from openpyxl.drawing.image import Image as XLImage
from sqlalchemy import func 
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

app = Flask(__name__)
app.config.from_object(Config)
# File upload configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'pdf'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = app.config.get('APP_SESSION_COOKIE_SECURE', False)

# Enable CORS for frontend
CORS(app, supports_credentials=True)

# Initialize database
db.init_app(app)

# Create tables
with app.app_context():
    db.create_all()

# File number management
SEQUENCE_FILE = 'last_sequence.txt'
DAM_LEVEL_CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'instance', 'dam_level_cache.json')

ROLE_PERMISSIONS = {
    'admin': {
        'manage_users',
        'view_physchem', 'edit_physchem',
        'view_micro', 'edit_micro',
        'view_water_treatment', 'edit_water_treatment',
        'view_leave', 'edit_leave',
        'view_screen_data', 'edit_screen_data',
        'view_foi', 'edit_foi'
    },
    'editor': {
        'view_physchem', 'edit_physchem',
        'view_micro', 'edit_micro',
        'view_water_treatment', 'edit_water_treatment',
        'view_leave', 'edit_leave',
        'view_screen_data', 'edit_screen_data',
        'view_foi', 'edit_foi'
    },
    'viewer': {
        'view_physchem',
        'view_micro',
        'view_water_treatment',
        'view_leave',
        'view_screen_data',
        'view_foi'
    },
    'phychemanalyst': {
        'view_physchem', 'edit_physchem',
        'view_leave', 'edit_leave',
        'view_screen_data', 'edit_screen_data',
        'view_foi', 'edit_foi'
    },
    'microanalyst': {
        'view_micro', 'edit_micro',
        'view_leave', 'edit_leave',
        'view_screen_data', 'edit_screen_data',
        'view_foi', 'edit_foi'
    },
    'guest': {
        'view_physchem',
        'view_micro',
        'view_water_treatment',
        'view_screen_data'
    },
    'viewonly': {
        'view_physchem',
        'view_micro',
        'view_water_treatment',
        'view_leave',
        'view_screen_data',
        'view_foi'
    }
}

SECTION_PRIVILEGE_MAP = {
    'physical-chemical': ('view_physchem', 'edit_physchem'),
    'microbiological': ('view_micro', 'edit_micro'),
    'water-treatment': ('view_water_treatment', 'edit_water_treatment'),
    'employee-leave': ('view_leave', 'edit_leave'),
    'foi-request': ('view_foi', 'edit_foi'),
    'screen-data': ('view_screen_data', 'edit_screen_data')
}


def _get_permissions_for_role(role: str):
    return ROLE_PERMISSIONS.get((role or '').lower(), set())


def _serialize_user(user: AppUser):
    permissions = _get_permissions_for_role(user.role)
    section_access = {
        section_id: {
            'view': view_permission in permissions,
            'edit': edit_permission in permissions
        }
        for section_id, (view_permission, edit_permission) in SECTION_PRIVILEGE_MAP.items()
    }

    return {
        **user.to_dict(),
        'permissions': sorted(list(permissions)),
        'sectionAccess': section_access
    }


def _ensure_default_admin_user():
    default_username = (app.config.get('APP_DEFAULT_ADMIN_USERNAME') or 'admin').strip()
    default_password = (app.config.get('APP_DEFAULT_ADMIN_PASSWORD') or 'admin123').strip()

    if not default_username or not default_password:
        return

    existing = AppUser.query.filter_by(username=default_username).first()
    if existing:
        return

    admin_user = AppUser(
        username=default_username,
        password_hash=generate_password_hash(default_password),
        role='admin',
        is_active=True
    )
    db.session.add(admin_user)
    db.session.commit()


def _get_current_user():
    user_id = session.get('user_id')
    if not user_id:
        return None
    user = AppUser.query.get(user_id)
    if not user or not user.is_active:
        session.clear()
        return None
    return user


def _get_effective_user():
    user = _get_current_user()
    if user:
        return user

    fallback_admin = AppUser.query.filter_by(is_active=True, role='admin').order_by(AppUser.id.asc()).first()
    if fallback_admin:
        return fallback_admin

    return AppUser.query.filter_by(is_active=True).order_by(AppUser.id.asc()).first()


def _is_auth_required() -> bool:
    return bool(app.config.get('APP_AUTH_REQUIRED', False))


def _require_permission(permission_name: str):
    if not _is_auth_required():
        return None

    current_user = g.get('current_user') or _get_current_user()
    if not current_user:
        return jsonify({'error': 'Authentication required'}), 401

    permissions = _get_permissions_for_role(current_user.role)
    if permission_name not in permissions:
        return jsonify({'error': 'Forbidden'}), 403

    return None


@app.before_request
def enforce_api_authentication():
    current_user = _get_current_user() if _is_auth_required() else _get_effective_user()
    g.current_user = current_user

    if not _is_auth_required():
        return None

    if request.method == 'OPTIONS':
        return None

    if not request.path.startswith('/api/'):
        return None

    public_paths = {
        '/api/health',
        '/api/auth/login'
    }
    if request.path in public_paths:
        return None

    if not current_user:
        return jsonify({'error': 'Authentication required'}), 401

    return None


with app.app_context():
    _ensure_default_admin_user()


def _hour_label_for_target(now: datetime, delay_minutes: int) -> str:
    expected = now - timedelta(minutes=delay_minutes)
    hour = expected.strftime('%I').lstrip('0') or '0'
    return f"{hour}:00 {expected.strftime('%p')}"


def _normalize_hour_header(raw_text: str) -> str:
    text = ' '.join((raw_text or '').split())
    match = re.search(r'(\d{1,2}):(\d{2})\s*([AP]M)', text, flags=re.IGNORECASE)
    if not match:
        return ''
    hour = str(int(match.group(1)))
    minute = match.group(2)
    meridiem = match.group(3).upper()
    return f"{hour}:{minute} {meridiem}"


def _parse_numeric(cell_text: str):
    text = (cell_text or '').strip()
    if not text:
        return None
    match = re.search(r'-?\d[\d,]*(?:\.\d+)?', text)
    if not match:
        return None
    return float(match.group(0).replace(',', ''))


def _target_slot_datetime(now: datetime, delay_minutes: int) -> datetime:
    expected = now - timedelta(minutes=delay_minutes)
    return expected.replace(minute=0, second=0, microsecond=0)


def _persist_and_get_recent_dam_snapshots(slot_datetime: datetime, target_hour_label: str, dam_level):
    try:
        if dam_level is not None:
            existing = DamLevelSnapshot.query.filter_by(slot_datetime=slot_datetime).first()
            if existing:
                existing.dam_level = float(dam_level)
                existing.target_hour = target_hour_label
            else:
                db.session.add(DamLevelSnapshot(
                    slot_datetime=slot_datetime,
                    target_hour=target_hour_label,
                    dam_level=float(dam_level)
                ))

        db.session.commit()
    except Exception:
        db.session.rollback()
        app.logger.exception('Failed to persist dam level snapshot')

    return DamLevelSnapshot.query.order_by(DamLevelSnapshot.slot_datetime.desc()).limit(4).all()


def _persist_and_get_recent_turbidity_snapshots(slot_datetime: datetime, target_hour_label: str, turbidity):
    try:
        if turbidity is not None:
            existing = TurbiditySnapshot.query.filter_by(slot_datetime=slot_datetime).first()
            if existing:
                existing.turbidity = float(turbidity)
                existing.target_hour = target_hour_label
            else:
                db.session.add(TurbiditySnapshot(
                    slot_datetime=slot_datetime,
                    target_hour=target_hour_label,
                    turbidity=float(turbidity)
                ))

        db.session.commit()
    except Exception:
        db.session.rollback()
        app.logger.exception('Failed to persist turbidity snapshot')

    return TurbiditySnapshot.query.order_by(TurbiditySnapshot.slot_datetime.desc()).limit(4).all()


def _load_dam_history():
    payload = _load_dam_cache_payload()
    history = payload.get('history', []) if isinstance(payload, dict) else []
    return history if isinstance(history, list) else []


def _build_screen_data_history(start_date: datetime = None, end_date: datetime = None):
    dam_query = DamLevelSnapshot.query
    turbidity_query = TurbiditySnapshot.query

    if start_date:
        dam_query = dam_query.filter(DamLevelSnapshot.slot_datetime >= start_date)
        turbidity_query = turbidity_query.filter(TurbiditySnapshot.slot_datetime >= start_date)

    if end_date:
        dam_query = dam_query.filter(DamLevelSnapshot.slot_datetime < end_date)
        turbidity_query = turbidity_query.filter(TurbiditySnapshot.slot_datetime < end_date)

    dam_rows = dam_query.order_by(DamLevelSnapshot.slot_datetime.asc()).all()
    turbidity_rows = turbidity_query.order_by(TurbiditySnapshot.slot_datetime.asc()).all()

    merged_by_slot = {}

    for row in dam_rows:
        key = row.slot_datetime.isoformat()
        merged_by_slot[key] = {
            'slotDatetime': row.slot_datetime.isoformat(),
            'date': row.slot_datetime.strftime('%Y-%m-%d'),
            'time': row.slot_datetime.strftime('%I:%M %p').lstrip('0'),
            'damLevel': float(row.dam_level) if row.dam_level is not None else None,
            'turbidity': None
        }

    for row in turbidity_rows:
        key = row.slot_datetime.isoformat()
        entry = merged_by_slot.get(key)
        if not entry:
            entry = {
                'slotDatetime': row.slot_datetime.isoformat(),
                'date': row.slot_datetime.strftime('%Y-%m-%d'),
                'time': row.slot_datetime.strftime('%I:%M %p').lstrip('0'),
                'damLevel': None,
                'turbidity': None
            }
            merged_by_slot[key] = entry

        entry['turbidity'] = float(row.turbidity) if row.turbidity is not None else None

    ordered_entries = sorted(merged_by_slot.values(), key=lambda item: item['slotDatetime'])
    grouped = defaultdict(list)

    for entry in ordered_entries:
        grouped[entry['date']].append(entry)

    return [
        {
            'date': date_key,
            'entries': grouped[date_key]
        }
        for date_key in sorted(grouped.keys())
    ]


def _build_missing_screen_data_hours(start_date: datetime = None, end_date: datetime = None):
    dam_query = DamLevelSnapshot.query
    turbidity_query = TurbiditySnapshot.query

    if start_date:
        dam_query = dam_query.filter(DamLevelSnapshot.slot_datetime >= start_date)
        turbidity_query = turbidity_query.filter(TurbiditySnapshot.slot_datetime >= start_date)

    if end_date:
        dam_query = dam_query.filter(DamLevelSnapshot.slot_datetime < end_date)
        turbidity_query = turbidity_query.filter(TurbiditySnapshot.slot_datetime < end_date)

    dam_rows = dam_query.order_by(DamLevelSnapshot.slot_datetime.asc()).all()
    turbidity_rows = turbidity_query.order_by(TurbiditySnapshot.slot_datetime.asc()).all()

    dam_by_slot = {}
    turbidity_by_slot = {}

    for row in dam_rows:
        slot = row.slot_datetime.replace(minute=0, second=0, microsecond=0)
        dam_by_slot[slot] = float(row.dam_level) if row.dam_level is not None else None

    for row in turbidity_rows:
        slot = row.slot_datetime.replace(minute=0, second=0, microsecond=0)
        turbidity_by_slot[slot] = float(row.turbidity) if row.turbidity is not None else None

    all_slots = sorted(set(list(dam_by_slot.keys()) + list(turbidity_by_slot.keys())))

    if start_date is not None:
        scan_start = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    elif all_slots:
        first_slot = all_slots[0]
        scan_start = first_slot.replace(hour=0, minute=0, second=0, microsecond=0)
    else:
        return {
            'totalMissingHours': 0,
            'groups': []
        }

    if end_date is not None:
        scan_end = end_date
    elif all_slots:
        last_slot = all_slots[-1]
        scan_end = last_slot.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)
    else:
        scan_end = scan_start + timedelta(days=1)

    grouped = defaultdict(list)
    cursor = scan_start

    while cursor < scan_end:
        dam_value = dam_by_slot.get(cursor)
        turbidity_value = turbidity_by_slot.get(cursor)

        dam_missing = dam_value is None or dam_value == 0
        turbidity_missing = turbidity_value is None or turbidity_value == 0

        if dam_missing and turbidity_missing:
            date_key = cursor.strftime('%Y-%m-%d')
            grouped[date_key].append({
                'slotDatetime': cursor.isoformat(),
                'time': cursor.strftime('%I:%M %p').lstrip('0'),
                'damLevel': dam_value,
                'turbidity': turbidity_value
            })

        cursor += timedelta(hours=1)

    groups = [
        {
            'date': date_key,
            'entries': grouped[date_key]
        }
        for date_key in sorted(grouped.keys())
    ]

    return {
        'totalMissingHours': sum(len(group['entries']) for group in groups),
        'groups': groups
    }


def _upsert_manual_screen_data_entries(entries):
    if not isinstance(entries, list):
        raise ValueError('Entries must be an array.')

    saved_count = 0

    for item in entries:
        if not isinstance(item, dict):
            continue

        slot_value = (item.get('slotDatetime') or '').strip()
        if not slot_value:
            continue

        try:
            slot_datetime = datetime.fromisoformat(slot_value)
        except ValueError as exc:
            raise ValueError(f'Invalid slotDatetime: {slot_value}') from exc

        slot_datetime = slot_datetime.replace(minute=0, second=0, microsecond=0)

        dam_level = item.get('damLevel')
        turbidity = item.get('turbidity')

        parsed_dam_level = None
        parsed_turbidity = None

        if dam_level is not None:
            try:
                parsed_dam_level = float(dam_level)
            except (TypeError, ValueError) as exc:
                raise ValueError(f'Invalid damLevel value for {slot_value}') from exc

        if turbidity is not None:
            try:
                parsed_turbidity = float(turbidity)
            except (TypeError, ValueError) as exc:
                raise ValueError(f'Invalid turbidity value for {slot_value}') from exc

        if parsed_dam_level is None and parsed_turbidity is None:
            continue

        hour = slot_datetime.strftime('%I').lstrip('0') or '0'
        target_hour_label = f"{hour}:00 {slot_datetime.strftime('%p')}"

        if parsed_dam_level is not None:
            dam_row = DamLevelSnapshot.query.filter_by(slot_datetime=slot_datetime).first()
            if dam_row:
                dam_row.dam_level = parsed_dam_level
                dam_row.target_hour = target_hour_label
            else:
                db.session.add(DamLevelSnapshot(
                    slot_datetime=slot_datetime,
                    target_hour=target_hour_label,
                    dam_level=parsed_dam_level
                ))

        if parsed_turbidity is not None:
            turbidity_row = TurbiditySnapshot.query.filter_by(slot_datetime=slot_datetime).first()
            if turbidity_row:
                turbidity_row.turbidity = parsed_turbidity
                turbidity_row.target_hour = target_hour_label
            else:
                db.session.add(TurbiditySnapshot(
                    slot_datetime=slot_datetime,
                    target_hour=target_hour_label,
                    turbidity=parsed_turbidity
                ))

        saved_count += 1

    db.session.commit()
    return saved_count


def _get_treatment_activity_metrics(threshold: float = 5.0, reference_time: datetime = None):
    now = reference_time or datetime.now()

    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if month_start.month == 12:
        next_month_start = month_start.replace(year=month_start.year + 1, month=1)
    else:
        next_month_start = month_start.replace(month=month_start.month + 1)

    last_active_row = (
        TurbiditySnapshot.query
        .filter(
            TurbiditySnapshot.turbidity.isnot(None),
            TurbiditySnapshot.turbidity > threshold,
            TurbiditySnapshot.slot_datetime <= now
        )
        .order_by(TurbiditySnapshot.slot_datetime.desc())
        .first()
    )

    total_treatment_hours_month = (
        TurbiditySnapshot.query
        .filter(
            TurbiditySnapshot.turbidity.isnot(None),
            TurbiditySnapshot.turbidity > threshold,
            TurbiditySnapshot.slot_datetime >= month_start,
            TurbiditySnapshot.slot_datetime < next_month_start
        )
        .count()
    )

    last_active_treatment = None
    if last_active_row and last_active_row.slot_datetime:
        last_active_treatment = last_active_row.slot_datetime.strftime('%Y-%m-%d %H:%M')

    return last_active_treatment, total_treatment_hours_month


def _build_screen_data_fallback_payload(scrape_error: str = None):
    computed_last_active_treatment, total_treatment_hours_month = _get_treatment_activity_metrics()
    manual_last_active_treatment = _get_last_active_dosing()
    last_active_treatment = manual_last_active_treatment or computed_last_active_treatment

    payload = {
        'target_hour': None,
        'target_column': None,
        'turbidity': None,
        'previous_turbidity': None,
        'turbidity_1_hour_prior': None,
        'turbidity_2_hours_prior': None,
        'turbidity_3_hours_prior': None,
        'current_dam_level': None,
        'previous_dam_level': None,
        'dam_level_1_hour_prior': None,
        'dam_level_2_hours_prior': None,
        'dam_level_3_hours_prior': None,
        'old_res_status': None,
        'old_res_big_tank_level': None,
        'tank_a_level': None,
        'tank_b_level': None,
        'tank_cd_level': None,
        'current_operator': None,
        'last_active_dosing': last_active_treatment,
        'total_treatment_hours_month': total_treatment_hours_month,
        'reserved_metric': _get_last_chlorine_tank_change(),
        'fetched_at': datetime.now().isoformat()
    }

    if scrape_error:
        payload['scrape_error'] = scrape_error

    return payload


def _load_dam_cache_payload():
    if not os.path.exists(DAM_LEVEL_CACHE_FILE):
        return {}
    try:
        import json
        with open(DAM_LEVEL_CACHE_FILE, 'r', encoding='utf-8') as file_handle:
            payload = json.load(file_handle)
        return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}


def _save_dam_cache_payload(payload):
    try:
        import json
        os.makedirs(os.path.dirname(DAM_LEVEL_CACHE_FILE), exist_ok=True)
        cache_payload = payload if isinstance(payload, dict) else {}
        existing_payload = _load_dam_cache_payload()
        if (
            isinstance(existing_payload, dict)
            and existing_payload.get('last_chlorine_tank_change')
            and not cache_payload.get('last_chlorine_tank_change')
        ):
            cache_payload['last_chlorine_tank_change'] = existing_payload.get('last_chlorine_tank_change')
        if (
            isinstance(existing_payload, dict)
            and existing_payload.get('last_active_dosing')
            and not cache_payload.get('last_active_dosing')
        ):
            cache_payload['last_active_dosing'] = existing_payload.get('last_active_dosing')
        history = cache_payload.get('history', [])
        if isinstance(history, list):
            cache_payload['history'] = history[-50:]
        with open(DAM_LEVEL_CACHE_FILE, 'w', encoding='utf-8') as file_handle:
            json.dump(cache_payload, file_handle)
    except Exception:
        pass


def _get_last_chlorine_tank_change():
    payload = _load_dam_cache_payload()
    value = payload.get('last_chlorine_tank_change') if isinstance(payload, dict) else None
    return value if isinstance(value, str) and value.strip() else None


def _set_last_chlorine_tank_change(date_value: str):
    payload = _load_dam_cache_payload()
    if not isinstance(payload, dict):
        payload = {}

    normalized = (date_value or '').strip()
    if normalized:
        payload['last_chlorine_tank_change'] = normalized
    else:
        payload.pop('last_chlorine_tank_change', None)

    _save_dam_cache_payload(payload)
    return payload.get('last_chlorine_tank_change')


def _get_last_active_dosing():
    payload = _load_dam_cache_payload()
    value = payload.get('last_active_dosing') if isinstance(payload, dict) else None
    return value if isinstance(value, str) and value.strip() else None


def _set_last_active_dosing(value: str):
    payload = _load_dam_cache_payload()
    if not isinstance(payload, dict):
        payload = {}

    normalized = (value or '').strip()
    if normalized:
        payload['last_active_dosing'] = normalized
    else:
        payload.pop('last_active_dosing', None)

    _save_dam_cache_payload(payload)
    return payload.get('last_active_dosing')


def _get_previous_from_last_displayed(cache_payload, current_target_hour: str):
    if not isinstance(cache_payload, dict):
        return None

    last_displayed = cache_payload.get('last_displayed_current_dam')
    last_target_hour = cache_payload.get('last_displayed_target_hour')
    last_fetched_at = cache_payload.get('last_displayed_fetched_at')

    if last_displayed is None or not last_target_hour or not last_fetched_at:
        return None
    if last_target_hour == current_target_hour:
        return None

    try:
        timestamp = datetime.fromisoformat(last_fetched_at)
        if (datetime.now() - timestamp) > timedelta(hours=18):
            return None
        return float(last_displayed)
    except Exception:
        return None


def _find_previous_from_history(history, current_target_hour: str):
    now = datetime.now()
    for item in reversed(history):
        dam = item.get('dam_level')
        target_hour = item.get('target_hour')
        fetched_at = item.get('fetched_at')
        if dam is None or target_hour == current_target_hour or not fetched_at:
            continue
        try:
            timestamp = datetime.fromisoformat(fetched_at)
            if (now - timestamp) > timedelta(hours=18):
                continue
            return float(dam)
        except Exception:
            continue
    return None


async def _scrape_screen_data_live():
    try:
        from playwright.async_api import async_playwright
    except ImportError as exc:
        raise RuntimeError("Playwright is not installed in backend environment.") from exc

    login_url = app.config.get('MONITORING_LOGIN_URL')
    username = app.config.get('MONITORING_USERNAME')
    password = app.config.get('MONITORING_PASSWORD')
    headless = app.config.get('MONITORING_HEADLESS', True)
    delay_minutes = app.config.get('MONITORING_ENTRY_DELAY_MINUTES', 12)

    if not username or not password:
        raise RuntimeError("Monitoring credentials are missing. Set MONITORING_USERNAME and MONITORING_PASSWORD.")

    table_xpath = '/html/body/table/tbody/tr[2]/th/table/tbody/tr/td[2]/table[3]/tbody'
    target_hour_label = _hour_label_for_target(datetime.now(), delay_minutes)
    dam_cache_payload = _load_dam_cache_payload()

    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=headless)
        context = await browser.new_context()
        page = await context.new_page()
        page.set_default_timeout(45000)

        try:
            await page.goto(login_url, timeout=45000)
            await page.wait_for_selector('input[name="username"]', state='visible')
            await page.fill('input[name="username"]', username)
            await page.fill('input[name="password"]', password)
            await page.press('input[name="password"]', 'Enter')
            await page.wait_for_load_state('domcontentloaded', timeout=30000)
            await page.wait_for_selector(f'xpath={table_xpath}/tr[1]', state='visible', timeout=30000)

            header_cells = page.locator(f'xpath={table_xpath}/tr[1]/td[position() >= 4 and position() <= 11]')
            header_texts = await header_cells.all_inner_texts()
            header_map = {}
            for idx, text in enumerate(header_texts, start=4):
                normalized = _normalize_hour_header(text)
                if normalized:
                    header_map[normalized] = idx

            target_column = header_map.get(target_hour_label)
            if target_column is None:
                raise RuntimeError(f"Target hour {target_hour_label} is not available in shift headers.")

            async def get_row_cell_text(row_label: str, column_index: int) -> str:
                row_locator = page.locator(
                    f'xpath={table_xpath}/tr[td[1][contains(normalize-space(.), "{row_label}")]]'
                ).first
                if await row_locator.count() == 0:
                    return ''
                cell_locator = row_locator.locator(f'xpath=./td[{column_index}]').first
                if await cell_locator.count() == 0:
                    return ''
                return (await cell_locator.inner_text()).strip()

            async def get_tank_cell_text(phase_label: str, tank_label: str, column_index: int) -> str:
                row_locator = page.locator(
                    f'xpath={table_xpath}/tr[td[2][contains(normalize-space(.), "{phase_label}")] and td[3][normalize-space(.)="{tank_label}"]]'
                ).first
                if await row_locator.count() == 0:
                    row_locator = page.locator(
                        f'xpath={table_xpath}/tr[td[2][contains(normalize-space(.), "{phase_label}")] and td[3][contains(normalize-space(.), "{tank_label}")]]'
                    ).first
                    if await row_locator.count() == 0:
                        return ''
                cell_locator = row_locator.locator(f'xpath=./td[{column_index}]').first
                if await cell_locator.count() == 0:
                    return ''
                return (await cell_locator.inner_text()).strip()

            current_dam_text = await get_row_cell_text('Dam Level', target_column)
            current_dam_value = _parse_numeric(current_dam_text)

            previous_dam_value = None
            for col in range(target_column - 1, 3, -1):
                left_text = await get_row_cell_text('Dam Level', col)
                parsed = _parse_numeric(left_text)
                if parsed is not None:
                    previous_dam_value = parsed
                    break

            dam_level_1_hour_prior = None
            one_hour_column = target_column - 1
            if one_hour_column >= 4:
                one_hour_text = await get_row_cell_text('Dam Level', one_hour_column)
                dam_level_1_hour_prior = _parse_numeric(one_hour_text)

            dam_level_2_hours_prior = None
            two_hour_column = target_column - 2
            if two_hour_column >= 4:
                two_hour_text = await get_row_cell_text('Dam Level', two_hour_column)
                dam_level_2_hours_prior = _parse_numeric(two_hour_text)

            dam_level_3_hours_prior = None
            three_hour_column = target_column - 3
            if three_hour_column >= 4:
                three_hour_text = await get_row_cell_text('Dam Level', three_hour_column)
                dam_level_3_hours_prior = _parse_numeric(three_hour_text)

            turbidity_text = await get_row_cell_text('Turbidity', target_column)
            turbidity_value = _parse_numeric(turbidity_text)

            previous_turbidity_value = None
            for col in range(target_column - 1, 3, -1):
                left_text = await get_row_cell_text('Turbidity', col)
                parsed = _parse_numeric(left_text)
                if parsed is not None:
                    previous_turbidity_value = parsed
                    break

            turbidity_1_hour_prior = None
            if one_hour_column >= 4:
                one_hour_turbidity_text = await get_row_cell_text('Turbidity', one_hour_column)
                turbidity_1_hour_prior = _parse_numeric(one_hour_turbidity_text)

            turbidity_2_hours_prior = None
            if two_hour_column >= 4:
                two_hour_turbidity_text = await get_row_cell_text('Turbidity', two_hour_column)
                turbidity_2_hours_prior = _parse_numeric(two_hour_turbidity_text)

            turbidity_3_hours_prior = None
            if three_hour_column >= 4:
                three_hour_turbidity_text = await get_row_cell_text('Turbidity', three_hour_column)
                turbidity_3_hours_prior = _parse_numeric(three_hour_turbidity_text)

            old_res_status = await get_row_cell_text('Old Reservoir P3 Status', target_column)
            old_res_big_tank_level = None
            for row_label in [
                'Old Reservoir P3 Big Tank Water Level',
                'Old Reservoir P3 Big Tank Level',
                'Old Reservoir Big Tank Water Level',
                'Old Reservoir Big Tank Level'
            ]:
                candidate_text = await get_row_cell_text(row_label, target_column)
                candidate_value = _parse_numeric(candidate_text)
                if candidate_value is not None:
                    old_res_big_tank_level = candidate_value
                    break

            tank_a_level = _parse_numeric(await get_tank_cell_text('Phase 1', 'A', target_column))
            if tank_a_level is None:
                for row_label in [
                    'Tank Water Level Phase 1 A',
                    'Tank Water Level - Phase 1 A',
                    'Tank Water Level Phase 1',
                    'Tank Water Level A'
                ]:
                    candidate_text = await get_row_cell_text(row_label, target_column)
                    candidate_value = _parse_numeric(candidate_text)
                    if candidate_value is not None:
                        tank_a_level = candidate_value
                        break

            tank_b_level = _parse_numeric(await get_tank_cell_text('Phase 1', 'B', target_column))
            if tank_b_level is None:
                for row_label in [
                    'Tank Water Level Phase 1 B',
                    'Tank Water Level - Phase 1 B',
                    'Tank Water Level Phase 2 B',
                    'Tank Water Level - Phase 2 B',
                    'Tank Water Level Phase 2',
                    'Tank Water Level B'
                ]:
                    candidate_text = await get_row_cell_text(row_label, target_column)
                    candidate_value = _parse_numeric(candidate_text)
                    if candidate_value is not None:
                        tank_b_level = candidate_value
                        break

            tank_c_level = _parse_numeric(await get_tank_cell_text('Phase 2', 'C', target_column))
            tank_d_level = _parse_numeric(await get_tank_cell_text('Phase 2', 'D', target_column))
            tank_cd_candidates = [value for value in [tank_c_level, tank_d_level] if value is not None]
            tank_cd_level = (sum(tank_cd_candidates) / len(tank_cd_candidates)) if tank_cd_candidates else None
            if tank_cd_level is None:
                for row_label in [
                    'Tank Water Level Phase 2 C',
                    'Tank Water Level - Phase 2 C',
                    'Tank Water Level Phase 2 D',
                    'Tank Water Level - Phase 2 D',
                    'Tank Water Level Phase 3 C & D',
                    'Tank Water Level Phase 3 C&D',
                    'Tank Water Level - Phase 3 C & D',
                    'Tank Water Level Phase 3',
                    'Tank Water Level C & D',
                    'Tank Water Level C&D'
                ]:
                    candidate_text = await get_row_cell_text(row_label, target_column)
                    candidate_value = _parse_numeric(candidate_text)
                    if candidate_value is not None:
                        tank_cd_level = candidate_value
                        break
            current_operator = await get_row_cell_text('Encoded By', target_column)

            target_slot_datetime = _target_slot_datetime(datetime.now(), delay_minutes)
            recent_dam_snapshots = _persist_and_get_recent_dam_snapshots(
                target_slot_datetime,
                target_hour_label,
                current_dam_value
            )
            recent_turbidity_snapshots = _persist_and_get_recent_turbidity_snapshots(
                target_slot_datetime,
                target_hour_label,
                turbidity_value
            )

            snapshot_offset = 1 if current_dam_value is not None else 0

            if len(recent_dam_snapshots) > snapshot_offset and recent_dam_snapshots[snapshot_offset].dam_level is not None:
                previous_dam_value = float(recent_dam_snapshots[snapshot_offset].dam_level)
                dam_level_1_hour_prior = float(recent_dam_snapshots[snapshot_offset].dam_level)

            if len(recent_dam_snapshots) > snapshot_offset + 1 and recent_dam_snapshots[snapshot_offset + 1].dam_level is not None:
                dam_level_2_hours_prior = float(recent_dam_snapshots[snapshot_offset + 1].dam_level)

            if len(recent_dam_snapshots) > snapshot_offset + 2 and recent_dam_snapshots[snapshot_offset + 2].dam_level is not None:
                dam_level_3_hours_prior = float(recent_dam_snapshots[snapshot_offset + 2].dam_level)

            turbidity_snapshot_offset = 1 if turbidity_value is not None else 0

            if len(recent_turbidity_snapshots) > turbidity_snapshot_offset and recent_turbidity_snapshots[turbidity_snapshot_offset].turbidity is not None:
                previous_turbidity_value = float(recent_turbidity_snapshots[turbidity_snapshot_offset].turbidity)
                turbidity_1_hour_prior = float(recent_turbidity_snapshots[turbidity_snapshot_offset].turbidity)

            if len(recent_turbidity_snapshots) > turbidity_snapshot_offset + 1 and recent_turbidity_snapshots[turbidity_snapshot_offset + 1].turbidity is not None:
                turbidity_2_hours_prior = float(recent_turbidity_snapshots[turbidity_snapshot_offset + 1].turbidity)

            if len(recent_turbidity_snapshots) > turbidity_snapshot_offset + 2 and recent_turbidity_snapshots[turbidity_snapshot_offset + 2].turbidity is not None:
                turbidity_3_hours_prior = float(recent_turbidity_snapshots[turbidity_snapshot_offset + 2].turbidity)

            if current_dam_value is not None:
                dam_cache_payload['last_displayed_current_dam'] = current_dam_value
                dam_cache_payload['last_displayed_target_hour'] = target_hour_label
                dam_cache_payload['last_displayed_fetched_at'] = datetime.now().isoformat()
                _save_dam_cache_payload(dam_cache_payload)

            computed_last_active_treatment, total_treatment_hours_month = _get_treatment_activity_metrics()
            manual_last_active_treatment = _get_last_active_dosing()
            last_active_treatment = manual_last_active_treatment or computed_last_active_treatment

            return {
                'target_hour': target_hour_label,
                'target_column': target_column,
                'turbidity': turbidity_value,
                'previous_turbidity': previous_turbidity_value,
                'turbidity_1_hour_prior': turbidity_1_hour_prior,
                'turbidity_2_hours_prior': turbidity_2_hours_prior,
                'turbidity_3_hours_prior': turbidity_3_hours_prior,
                'current_dam_level': current_dam_value,
                'previous_dam_level': previous_dam_value,
                'dam_level_1_hour_prior': dam_level_1_hour_prior,
                'dam_level_2_hours_prior': dam_level_2_hours_prior,
                'dam_level_3_hours_prior': dam_level_3_hours_prior,
                'old_res_status': old_res_status or None,
                'old_res_big_tank_level': old_res_big_tank_level,
                'tank_a_level': tank_a_level,
                'tank_b_level': tank_b_level,
                'tank_cd_level': tank_cd_level,
                'current_operator': current_operator or None,
                'last_active_dosing': last_active_treatment,
                'total_treatment_hours_month': total_treatment_hours_month,
                'reserved_metric': dam_cache_payload.get('last_chlorine_tank_change'),
                'fetched_at': datetime.now().isoformat()
            }
        finally:
            await browser.close()


def _run_scrape_coroutine(coroutine_factory):
    try:
        return asyncio.run(coroutine_factory())
    except RuntimeError as error:
        if 'asyncio.run() cannot be called from a running event loop' not in str(error):
            raise
        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(coroutine_factory())
        finally:
            loop.close()

def create_water_treatment_excel_report(records, report_type, date_info):
    """Generate Excel report for water treatment records"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type} Report"
        
        # Header styling
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Title
        title_text = f"TURBIDITY MONITORING FOR THE MONTH OF {date_info.upper()}" if report_type.lower() == 'monthly' else f"TURBIDITY MONITORING - {date_info.upper()}"
        ws.merge_cells('A1:M1')
        title_cell = ws['A1']
        title_cell.value = title_text
        title_cell.font = Font(name='Calibri', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        
        # Headers - Row 3
        headers = [
            'DATE', 'DAM LEVEL', 'RAW WATER', 
            'CLARIFIED WATER 1', 'CLARIFIED WATER 2',
            'FILTERED WATER 1', 'FILTERED WATER 2',
            'PAC DOSAGE\n(L/min)', 'ALUM DOSAGE\n(g/m3)',
            'PAC CONSUMPTION\n(L/hr)', 'ALUM CONSUMPTION\n(BAGS/HR)', 'TREATMENT HOURS'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Data rows and consumption calculation
        # Data rows and consumption calculation
        row_num = 4
        total_alum_bags = 0
        total_pac_liters = 0
        total_treatment_hours = 0
        dam_levels = []
        
        for i, record in enumerate(records):
            # Calculate hours until next reading
            treatment_hours = 0
            pac_consumption = 0
            alum_consumption = 0
            
            # Only calculate if raw water > 5 NTU (treatment is needed)
            requires_treatment = record.raw_water_turbidity and record.raw_water_turbidity > 5
            
            if requires_treatment and i < len(records) - 1:
                next_record = records[i + 1]
                treatment_hours = (next_record.reading_datetime - record.reading_datetime).total_seconds() / 3600
                
                # PAC: L/min × 60 × hours = liters
                if record.pac_dosage:
                    pac_consumption = record.pac_dosage * 60 * treatment_hours
                    total_pac_liters += pac_consumption
                
                # Alum: % Pump × 0.06 × hours = bags
                if record.alum_dosage:
                    alum_consumption = record.alum_dosage * 0.06 * treatment_hours
                    total_alum_bags += alum_consumption
                
                total_treatment_hours += treatment_hours
            
            # Track dam levels for average
            if record.dam_level:
                dam_levels.append(record.dam_level)
            
            # Write data
            ws.cell(row=row_num, column=1, value=record.reading_datetime.strftime('%m/%d/%y %I:%M %p') if record.reading_datetime else '')
            ws.cell(row=row_num, column=2, value=record.dam_level if record.dam_level else '')
            ws.cell(row=row_num, column=3, value=record.raw_water_turbidity if record.raw_water_turbidity else '')
            ws.cell(row=row_num, column=4, value=record.clarified_water_phase1 if record.clarified_water_phase1 else '')
            ws.cell(row=row_num, column=5, value=record.clarified_water_phase2 if record.clarified_water_phase2 else '')
            ws.cell(row=row_num, column=6, value=record.filtered_water_phase1 if record.filtered_water_phase1 else '')
            ws.cell(row=row_num, column=7, value=record.filtered_water_phase2 if record.filtered_water_phase2 else '')
            ws.cell(row=row_num, column=8, value=record.pac_dosage if record.pac_dosage else '')
            ws.cell(row=row_num, column=9, value=record.alum_dosage if record.alum_dosage else '')
            
            # Only show consumption if treatment was required
            if requires_treatment:
                ws.cell(row=row_num, column=10, value=round(pac_consumption, 2) if pac_consumption > 0 else '')
                ws.cell(row=row_num, column=11, value=round(alum_consumption, 2) if alum_consumption > 0 else '')
                ws.cell(row=row_num, column=12, value=round(treatment_hours, 2) if treatment_hours > 0 else '')
            else:
                # If no treatment needed, show 0 or leave blank
                ws.cell(row=row_num, column=10, value='')
                ws.cell(row=row_num, column=11, value='')
                ws.cell(row=row_num, column=12, value='')
            
            # Add borders and alignment
            for col in range(1, 13):
                cell = ws.cell(row=row_num, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row_num += 1
        
        # Add summary rows for monthly reports
        if report_type.lower() == 'monthly' and len(records) > 0:
            row_num += 1
            avg_dam_level = sum(dam_levels) / len(dam_levels) if dam_levels else 0
            
            # Average Dam Level
            ws.cell(row=row_num, column=1, value="AVERAGE DAM LEVEL").font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=round(avg_dam_level, 2)).font = Font(bold=True)
            row_num += 1
            
            # Total Treatment Hours
            ws.cell(row=row_num, column=1, value="TOTAL TREATMENT HOURS").font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=round(total_treatment_hours, 2)).font = Font(bold=True)
            row_num += 1
            
            # Total PAC Consumption
            ws.cell(row=row_num, column=1, value="TOTAL PAC CONSUMPTION").font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=round(total_pac_liters, 2)).font = Font(bold=True)
            row_num += 1
            
            # Total Alum Consumption
            ws.cell(row=row_num, column=1, value="TOTAL ALUM CONSUMPTION").font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=round(total_alum_bags, 2)).font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 18
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 18
        
        # Save file
        filename = f"Water_Treatment_{report_type}_{date_info.replace(' ', '_').replace('/', '-')}.xlsx"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'water_treatment', filename)
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        wb.save(filepath)
        app.logger.info("Water treatment report created: %s", filename)
        app.logger.info("Water treatment totals alum_bags=%.2f pac_liters=%.2f treatment_hours=%.2f", total_alum_bags, total_pac_liters, total_treatment_hours)
        return filename
        
    except Exception as e:
        app.logger.exception("Error creating water treatment report")
        return None


def get_current_month():
    return datetime.now().strftime("%Y-%m")

def get_next_file_number(peek=True):
    """Get next file number, optionally incrementing"""
    current_month = get_current_month()
    
    # Read existing sequence
    last_month = None
    last_seq = 0
    
    if os.path.exists(SEQUENCE_FILE):
        try:
            with open(SEQUENCE_FILE, 'r') as f:
                data = f.read().strip().split(',')
                if len(data) == 2:
                    last_month, last_seq_str = data
                    last_seq = int(last_seq_str)
        except:
            pass
    
    # Reset if new month or start at 1 if no sequence
    if last_month != current_month:
        if peek:
            sequence = 1
        else:
            sequence = 1
            # Save the first number when incrementing in a new month
            with open(SEQUENCE_FILE, 'w') as f:
                f.write(f"{current_month},{sequence}")
    else:
        if peek:
            # Just return current sequence + 1 without saving
            sequence = last_seq + 1
        else:
            # Increment and save
            sequence = last_seq + 1
            with open(SEQUENCE_FILE, 'w') as f:
                f.write(f"{current_month},{sequence}")
    
    return str(sequence).zfill(3)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_excel_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}

def write_to_cell(sheet, cell_address, value):
    """Write to a cell, handling merged cells"""
    target_cell = sheet[cell_address]
    
    # Check if this cell is part of a merged range
    for merged_range in sheet.merged_cells.ranges:
        if target_cell.coordinate in merged_range:
            # Write to the top-left cell of the merged range
            top_left_cell = merged_range.start_cell
            sheet.cell(row=top_left_cell.row, column=top_left_cell.column).value = value
            return
    
    # If not merged, write directly
    target_cell.value = value

def set_sheet_to_a4(sheet):
    """Set sheet to print perfectly on A4 paper - fits on one page"""
    # Page setup
    sheet.page_setup.orientation = 'portrait'
    sheet.page_setup.paperSize = 9  # A4
    
    # Use scaling for better control
    sheet.page_setup.scale = 70  # Scale to 70% for micro (slightly smaller)
    sheet.page_setup.fitToHeight = False
    sheet.page_setup.fitToWidth = False
    
    # Print area - adjust based on sheet type
    if 'Micro' in sheet.title or 'micro' in sheet.title.lower():
        sheet.print_area = 'A1:S35'  # Smaller range for micro
    else:
        sheet.print_area = 'A1:S46'  # PhysChem range
    
    # Smaller margins
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.3
    sheet.page_margins.bottom = 0.2
    sheet.page_margins.header = 0.1
    sheet.page_margins.footer = 0.1
    
    # Center horizontally
    sheet.print_options.horizontalCentered = True
    sheet.print_options.gridLines = False
    sheet.print_options.gridLinesSet = True
    
    # Remove headers/footers
    sheet.oddHeader.left.text = ""
    sheet.oddHeader.center.text = ""
    sheet.oddHeader.right.text = ""
    sheet.oddFooter.left.text = ""
    sheet.oddFooter.center.text = ""
    sheet.oddFooter.right.text = ""

def create_physchem_excel(analysis, analyst_signature_scale=100, approver_signature_scale=100):
    """Generate Excel file from PhysChemAnalysis record using template"""
    try:
        template_path = 'Form.xlsx'
        
        # Load template
        if not os.path.exists(template_path):
            app.logger.error("PhysChem template not found: %s", template_path)
            return None
        
        wb = openpyxl.load_workbook(template_path)
        
        # Get the correct analyst sheet or use first sheet
        analyst_name = analysis.analyst if analysis.analyst else "Benjamin"
        sheet_name = analyst_name.split()[0] if analyst_name else "Benjamin"
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        ws.column_dimensions['N'].width = 20  # <-- ADD THIS LINE
        
        # Cell mapping
        cell_map = {
            "Client": "C9",
            "Source": "C10",
            "Location": "C11",
            "Date Collected": "F13",
            "Date Analyzed": "F14",
            "Date Submitted": "S13",
            "Collected By": "S14",
            "File #": "Q10",
            "O.R. #": "Q11",
            "Analyst": "B42",
            "pH": "F20",
            "Turbidity": "F21",
            "Color": "F22",
            "Iron": "F23",
            "Chloride": "F24",
            "Copper": "F25",
            "Chromium": "F26",
            "Manganese": "F27",
            "Total Hardness": "F28",
            "Sulfate": "F29",
            "Nitrate": "F30",
            "Nitrite": "F31",
            "Total Dissolved Solids": "F32"
        }
        
        # Fill in data using the helper function
        write_to_cell(ws, cell_map["Client"], analysis.client)
        write_to_cell(ws, cell_map["Source"], analysis.source)
        write_to_cell(ws, cell_map["Location"], analysis.location)
        write_to_cell(ws, cell_map["Date Collected"], str(analysis.date_collected) if analysis.date_collected else '')
        write_to_cell(ws, cell_map["Date Analyzed"], str(analysis.date_analyzed) if analysis.date_analyzed else '')
        write_to_cell(ws, cell_map["Date Submitted"], str(analysis.date_submitted) if analysis.date_submitted else '')
        write_to_cell(ws, cell_map["Collected By"], analysis.collected_by)
        write_to_cell(ws, cell_map["Analyst"], analysis.analyst)

# Format analyst name (make it bold and handle long names)
        if analysis.analyst:
            # Use smaller font for Manuel's long name
            if 'Manuel Benjamin Obsequio' in analysis.analyst:
                ws['B42'].font = Font(name='Calibri', size=11, bold=True)
            else:
                ws['B42'].font = Font(name='Calibri', size=12, bold=True)
            
            ws['B42'].alignment = Alignment(horizontal='center', vertical='center')

        analyst_titles = {
            'Benjamin A. Lasola Jr.': 'Senior Laboratory Technician',
            'Crispulo War Y. Indoc': 'Quality Control Officer',
            'Allan Mark R. Ong': 'Quality Control Officer',
            'Manuel Benjamin Obsequio': 'Quality Control Officer'
}

# Update the title in cell A46 (below the analyst signature)
        # Place the title in Row 43, Columns B:I (merged)
        if analysis.analyst and analysis.analyst in analyst_titles:
            analyst_title = analyst_titles[analysis.analyst]
            
            # Merge cells B43:I43
            ws.merge_cells('B43:I43')
            
            # Write the title
            ws['B43'] = analyst_title
            ws['B43'].font = Font(name='Calibri', size=12, bold=False) 
            ws['B43'].alignment = Alignment(horizontal='center', vertical='center')
            app.logger.debug("Updated analyst title to %s in B43:I43", analyst_title)
         
        write_to_cell(ws, cell_map["File #"], f"{analysis.file_prefix}{analysis.file_number}" if analysis.file_number else "")
        write_to_cell(ws, cell_map["O.R. #"], analysis.or_number)
        
        # Fill parameters
        write_to_cell(ws, cell_map["pH"], analysis.pH if analysis.pH is not None else '')
        write_to_cell(ws, cell_map["Turbidity"], analysis.turbidity if analysis.turbidity is not None else '')
        write_to_cell(ws, cell_map["Color"], analysis.color if analysis.color is not None else '')
        write_to_cell(ws, cell_map["Iron"], analysis.iron if analysis.iron is not None else '')
        write_to_cell(ws, cell_map["Chloride"], analysis.chloride if analysis.chloride is not None else '')
        write_to_cell(ws, cell_map["Copper"], analysis.copper if analysis.copper is not None else '')
        write_to_cell(ws, cell_map["Chromium"], analysis.chromium if analysis.chromium is not None else '')
        write_to_cell(ws, cell_map["Manganese"], analysis.manganese if analysis.manganese is not None else '')
        write_to_cell(ws, cell_map["Total Hardness"], analysis.total_hardness if analysis.total_hardness is not None else '')
        write_to_cell(ws, cell_map["Sulfate"], analysis.sulfate if analysis.sulfate is not None else '')
        write_to_cell(ws, cell_map["Nitrate"], analysis.nitrate if analysis.nitrate is not None else '')
        write_to_cell(ws, cell_map["Nitrite"], analysis.nitrite if analysis.nitrite is not None else '')
        write_to_cell(ws, cell_map["Total Dissolved Solids"], analysis.total_dissolved_solids if analysis.total_dissolved_solids is not None else '')
        
         # ADD SIGNATURE IMAGE
        write_to_cell(ws, cell_map["Total Dissolved Solids"], analysis.total_dissolved_solids if analysis.total_dissolved_solids is not None else '')
        
        # ADD SIGNATURE IMAGE
        try:
            # Map analyst names to signature files
            analyst_signatures = {
                'Benjamin A. Lasola Jr.': 'benjamin_signature.png',
                'Crispulo War Y. Indoc': 'crispulo_signature.png',
                'Allan Mark R. Ong': 'allan_signature.png',
                'Manuel Benjamin Obsequio': 'manuel_signature.png'
            }
            
            if analysis.analyst and analysis.analyst in analyst_signatures:
                sig_filename = analyst_signatures[analysis.analyst]
                sig_path = os.path.join(os.path.dirname(__file__), '..', 'signatures', sig_filename)
                sig_path = os.path.abspath(sig_path)
                analyst_scale = max(50, min(200, float(analyst_signature_scale or 100))) / 100.0
                
                if os.path.exists(sig_path):
                    img = XLImage(sig_path)
                    img.width = 200 * analyst_scale
                    img.height = 60 * analyst_scale
                    ws.add_image(img, 'E40')
                    app.logger.debug("Added analyst signature for %s", analysis.analyst)
                else:
                    app.logger.warning("Analyst signature file not found: %s", sig_path)
        except Exception as e:
            app.logger.exception("Error adding analyst signature to PhysChem Excel")

        # ADD SEAL IMAGE
        try:
            seal_path = os.path.join(os.path.dirname(__file__), '..', 'zcwd_seal.png')
            
            if os.path.exists(seal_path):
                seal_img = XLImage(seal_path)
                seal_img.width = 80
                seal_img.height = 80
                ws.add_image(seal_img, 'L1')
                app.logger.debug("Added ZCWD seal to PhysChem Excel")
            else:
                app.logger.warning("ZCWD seal not found at: %s", seal_path)
        except Exception as e:
            app.logger.exception("Error adding ZCWD seal to PhysChem Excel")

        try:
            eric_sig_path = os.path.join(os.path.dirname(__file__), '..', 'signatures', 'eric_signature.png')
            eric_sig_path = os.path.abspath(eric_sig_path)
            
            if os.path.exists(eric_sig_path):
                approver_scale = max(50, min(200, float(approver_signature_scale or 100))) / 100.0
                eric_img = XLImage(eric_sig_path)
                eric_img.width = 180 * approver_scale
                eric_img.height = 70 * approver_scale
                ws.add_image(eric_img, 'N40')
                app.logger.debug("Added Eric V. Salaritan signature to PhysChem Excel")
            else:
                app.logger.warning("Eric signature file not found: %s", eric_sig_path)
        except Exception as e:
            app.logger.exception("Error adding Eric signature to PhysChem Excel")

        # Generate filename
        client_clean = analysis.client.replace(" ", "_").replace("/", "-")[:30] if analysis.client else "Unknown"
        date_str = str(analysis.date_analyzed).replace("-", "") if analysis.date_analyzed else datetime.now().strftime("%Y%m%d")
        file_num = f"{analysis.file_prefix}{analysis.file_number}" if analysis.file_number else "NoNum"
        filename = f"{client_clean}_{date_str}_{file_num}.xlsx"
        
        set_sheet_to_a4(ws)

        # Save file
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'physchem', filename)
        wb.save(filepath)
        app.logger.info("PhysChem Excel file created: %s", filename)
        return filename
    except Exception as e:
        app.logger.exception("Error creating PhysChem Excel file")
        return None

def create_micro_excel(
    analysis,
    show_benjamin=False,
    show_eric=False,
    analyst_signature_scale=100,
    approver_signature_scale=100
):
    """Generate Excel file from MicrobiologicalAnalysis record using template"""
    try:
        template_path = 'MicroTemplate.xlsx'
        
        # Load template
        if not os.path.exists(template_path):
            app.logger.error("Micro template not found: %s", template_path)
            return None
        
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Fill in client information (based on your template)
        ws['B9'] = analysis.client or ''  # Client
        ws['B10'] = analysis.source or ''  # Source
        ws['B11'] = analysis.location or ''  # Location
        
        ws['B13'] = str(analysis.date_collected) if analysis.date_collected else ''  # Date Collected
        ws['B14'] = str(analysis.date_analyzed) if analysis.date_analyzed else ''  # Date Analyzed
        
        # Right side
        file_num_text = f"{analysis.file_prefix}{analysis.file_number}" if analysis.file_number else ""
        ws['F9'] = file_num_text  # FILE #
        ws['F10'] = analysis.or_number or ''  # O.R. #
        ws['F13'] = str(analysis.date_submitted) if analysis.date_submitted else ''  # Date Submitted
        ws['F14'] = analysis.collected_by or ''  # Collected By
        
        # Fill in test results (rows 19-22)
        # Fill in test results (rows 19-22)
        test_data = [
            {
                'row': 19,
                'value': analysis.total_coliform,
                'unit': 'MPN/100mL',
                'standard_value': 1.1
            },
            {
                'row': 20,
                'value': analysis.e_coli,
                'unit': 'MPN/100mL',
                'standard_value': 1.1
            },
            {
                'row': 21,
                'value': analysis.fecal_coliform,
                'unit': 'MPN/100mL',
                'standard_value': 1.1
            },
            {
                'row': 22,
                'value': analysis.heterotrophic_plate_count,
                'unit': 'CFU/mL',
                'standard_value': 500
            }
        ]

        for test in test_data:
            row = test['row']
            value = test['value']
            unit = test['unit']
            standard_value = test['standard_value']
            
            if value is not None:
                # Result with unit in column C (merged with D)
                result_text = f"{value} {unit}"
                ws[f'C{row}'] = result_text
                ws[f'C{row}'].font = Font(name='Calibri', size=10)
                ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Update remarks in column F
                is_positive = float(value) >= standard_value
                remark = 'POSITIVE' if is_positive else 'NEGATIVE'
                color = 'FF0000' if is_positive else '00FF00'
                
                ws[f'F{row}'] = remark
                ws[f'F{row}'].font = Font(name='Calibri', size=10, bold=True, color=color)
        
        # Add Benjamin's signature
        if show_benjamin:
            try:
                ben_sig_path = os.path.join(os.path.dirname(__file__), '..', 'signatures', 'benjamin_signature.png')
                if os.path.exists(ben_sig_path):
                    analyst_scale = max(50, min(200, float(analyst_signature_scale or 100))) / 100.0
                    ben_img = XLImage(ben_sig_path)
                    ben_img.width = 120 * analyst_scale
                    ben_img.height = 40 * analyst_scale
                    ben_img.anchor = 'B31'
                    ws.add_image(ben_img)
                    app.logger.debug("Added Benjamin signature to Micro Excel")
            except Exception as e:
                app.logger.exception("Error adding Benjamin signature to Micro Excel")
        else:
            app.logger.debug("Skipping Benjamin signature for Micro Excel")

        # Eric's signature - only if requested  
        if show_eric:
            try:
                eric_sig_path = os.path.join(os.path.dirname(__file__), '..', 'signatures', 'eric_signature.png')
                if os.path.exists(eric_sig_path):
                    approver_scale = max(50, min(200, float(approver_signature_scale or 100))) / 100.0
                    eric_img = XLImage(eric_sig_path)
                    eric_img.width = 120 * approver_scale
                    eric_img.height = 40 * approver_scale
                    eric_img.anchor = 'E31'
                    ws.add_image(eric_img)
                    app.logger.debug("Added Eric signature to Micro Excel")
            except Exception as e:
                app.logger.exception("Error adding Eric signature to Micro Excel")
        else:
            app.logger.debug("Skipping Eric signature for Micro Excel")
        
        
        # Generate filename
        client_clean = analysis.client.replace(" ", "_").replace("/", "-")[:30] if analysis.client else "Unknown"
        date_str = str(analysis.date_analyzed).replace("-", "") if analysis.date_analyzed else datetime.now().strftime("%Y%m%d")
        file_num = f"{analysis.file_prefix}{analysis.file_number}" if analysis.file_number else "NoNum"
        filename = f"MICRO_{client_clean}_{date_str}_{file_num}.xlsx"
        
        
        try:
            ws.page_setup.orientation = 'portrait'
            ws.page_setup.paperSize = 9  # A4
            ws.page_setup.fitToHeight = 1
            ws.page_setup.fitToWidth = 1
        except Exception as e:
            app.logger.warning("Could not set all Micro page properties: %s", e)
   
        
        # Save file
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'micro', filename)
        wb.save(filepath)
        app.logger.info("Micro Excel file created: %s", filename)
        return filename
    except Exception as e:
        app.logger.exception("Error creating Micro Excel file")
        return None

def convert_excel_to_pdf(excel_path):
    """Convert Excel file to PDF using LibreOffice"""
    try:
        import subprocess
        
        # Get directory and filename
        excel_dir = os.path.dirname(excel_path)
        excel_filename = os.path.basename(excel_path)
        pdf_filename = excel_filename.replace('.xlsx', '.pdf')
        pdf_path = os.path.join(excel_dir, pdf_filename)
        
        # Check if LibreOffice is available
        result = subprocess.run(
            ['libreoffice', '--version'],
            capture_output=True,
            timeout=5
        )
        
        if result.returncode != 0:
            app.logger.error("LibreOffice not found")
            return None
        
        # Convert Excel to PDF using LibreOffice headless
        subprocess.run([
            'libreoffice',
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', excel_dir,
            excel_path
        ], timeout=30)
        
        if os.path.exists(pdf_path):
            app.logger.info("PDF created successfully: %s", pdf_filename)
            return pdf_filename
        else:
            app.logger.error("PDF conversion failed for: %s", excel_path)
            return None
            
    except Exception as e:
        app.logger.exception("Error converting Excel to PDF: %s", excel_path)
        return None


@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    try:
        data = request.json or {}
        username = (data.get('username') or '').strip()
        password = str(data.get('password') or '')

        if _is_auth_required():
            if not username or not password:
                return jsonify({'error': 'Username and password are required'}), 400

            user = AppUser.query.filter_by(username=username, is_active=True).first()
            if not user or not check_password_hash(user.password_hash, password):
                return jsonify({'error': 'Invalid username or password'}), 401

            session['user_id'] = user.id
            return jsonify({'message': 'Login successful', 'user': _serialize_user(user)}), 200

        user = AppUser.query.filter_by(username=username, is_active=True).first() if username else _get_effective_user()
        if not user:
            return jsonify({'error': 'No active user available'}), 404

        session['user_id'] = user.id
        return jsonify({'message': 'Access granted', 'user': _serialize_user(user)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/auth/logout', methods=['POST'])
def auth_logout():
    session.clear()
    return jsonify({'message': 'Logout successful'}), 200


@app.route('/api/auth/me', methods=['GET'])
def auth_me():
    user = _get_current_user() if _is_auth_required() else _get_effective_user()
    if not user:
        if _is_auth_required():
            return jsonify({'error': 'Authentication required'}), 401
        return jsonify({'error': 'No active user available'}), 404
    return jsonify({'user': _serialize_user(user)}), 200


@app.route('/api/auth/users', methods=['GET'])
def list_auth_users():
    permission_error = _require_permission('manage_users')
    if permission_error:
        return permission_error

    users = AppUser.query.order_by(AppUser.username.asc()).all()
    return jsonify([entry.to_dict() for entry in users]), 200


@app.route('/api/auth/users', methods=['POST'])
def create_auth_user():
    permission_error = _require_permission('manage_users')
    if permission_error:
        return permission_error

    try:
        data = request.json or {}
        username = (data.get('username') or '').strip()
        password = (data.get('password') or '').strip()
        role = (data.get('role') or 'viewer').strip().lower()
        is_active = bool(data.get('isActive', True))

        if not username or not password:
            return jsonify({'error': 'Username and password are required'}), 400
        if role not in ROLE_PERMISSIONS:
            return jsonify({'error': 'Invalid role'}), 400
        if AppUser.query.filter_by(username=username).first():
            return jsonify({'error': 'Username already exists'}), 409

        new_user = AppUser(
            username=username,
            password_hash=generate_password_hash(password),
            role=role,
            is_active=is_active
        )
        db.session.add(new_user)
        db.session.commit()
        return jsonify({'message': 'User created successfully', 'user': new_user.to_dict()}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400


@app.route('/api/auth/users/<int:user_id>', methods=['PATCH'])
def update_auth_user(user_id):
    permission_error = _require_permission('manage_users')
    if permission_error:
        return permission_error

    target_user = AppUser.query.get_or_404(user_id)
    try:
        data = request.json or {}

        if 'role' in data:
            role = (data.get('role') or '').strip().lower()
            if role not in ROLE_PERMISSIONS:
                return jsonify({'error': 'Invalid role'}), 400
            target_user.role = role

        if 'isActive' in data:
            target_user.is_active = bool(data.get('isActive'))

        if 'password' in data and data.get('password'):
            target_user.password_hash = generate_password_hash(str(data.get('password')))

        db.session.commit()
        return jsonify({'message': 'User updated successfully', 'user': target_user.to_dict()}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400


# Routes for file number management
@app.route('/api/next-file-number', methods=['GET'])
def next_file_number():
    try:
        prefix = request.args.get('prefix', 'Pr')
        if prefix == 'Monitoring':
            return jsonify({'nextNumber': ''}), 200
        
        number = get_next_file_number(peek=True)
        return jsonify({'nextNumber': number}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/increment-file-number', methods=['POST'])
def increment_file_number():
    try:
        get_next_file_number(peek=False)
        return jsonify({'message': 'File number incremented'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# Health check
@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'ok', 'message': 'Backend is running!'})

# Physical Chemical Analysis Routes
@app.route('/api/physchem', methods=['POST'])
def create_physchem():
    try:
        data = request.json
        
        # Convert date strings to date objects
        date_collected = datetime.fromisoformat(data.get('dateCollected')).date() if data.get('dateCollected') else None
        date_analyzed = datetime.fromisoformat(data.get('dateAnalyzed')).date() if data.get('dateAnalyzed') else None
        date_submitted = datetime.fromisoformat(data.get('dateSubmitted')).date() if data.get('dateSubmitted') else None
        
        analyst_signature_scale = data.get('analystSignatureScale')
        approver_signature_scale = data.get('approverSignatureScale')

        analysis = PhysChemAnalysis(
            client=data.get('client'),
            source=data.get('source'),
            location=data.get('location'),
            date_collected=date_collected,
            date_analyzed=date_analyzed,
            date_submitted=date_submitted,
            file_prefix=data.get('filePrefix'),
            file_number=data.get('fileNumber'),
            or_number=data.get('orNumber'),
            collected_by=data.get('collectedBy'),
            analyst=data.get('analyst'),
            pH=float(data.get('pH')) if data.get('pH') else None,
            turbidity=float(data.get('turbidity')) if data.get('turbidity') else None,
            color=float(data.get('color')) if data.get('color') else None,
            total_dissolved_solids=float(data.get('totalDissolvedSolids')) if data.get('totalDissolvedSolids') else None,
            iron=float(data.get('iron')) if data.get('iron') else None,
            chloride=float(data.get('chloride')) if data.get('chloride') else None,
            copper=float(data.get('copper')) if data.get('copper') else None,
            chromium=float(data.get('chromium')) if data.get('chromium') else None,
            manganese=float(data.get('manganese')) if data.get('manganese') else None,
            total_hardness=float(data.get('totalHardness')) if data.get('totalHardness') else None,
            sulfate=float(data.get('sulfate')) if data.get('sulfate') else None,
            nitrate=float(data.get('nitrate')) if data.get('nitrate') else None,
            nitrite=float(data.get('nitrite')) if data.get('nitrite') else None
        )
        
        db.session.add(analysis)
        db.session.commit()
        
        # Generate Excel file
        excel_filename = create_physchem_excel(
            analysis,
            analyst_signature_scale=analyst_signature_scale,
            approver_signature_scale=approver_signature_scale
        )
        
        pdf_filename = None
        if excel_filename:
            excel_path = os.path.join(app.config['UPLOAD_FOLDER'], 'physchem', excel_filename)
            pdf_filename = convert_excel_to_pdf(excel_path)
        
        return jsonify({
            'message': 'PhysChem analysis saved successfully!',
            'id': analysis.id,
            'excel_file': excel_filename,
            'pdf_file': pdf_filename
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/physchem', methods=['GET'])
def get_all_physchem():
    try:
        analyses = PhysChemAnalysis.query.order_by(PhysChemAnalysis.created_at.desc()).all()
        return jsonify([analysis.to_dict() for analysis in analyses]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/physchem/<int:id>', methods=['GET'])
def get_physchem(id):
    try:
        analysis = PhysChemAnalysis.query.get_or_404(id)
        return jsonify(analysis.to_dict()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@app.route('/api/physchem/search', methods=['GET'])
def search_physchem():
    try:
        query = request.args.get('q', '')
        analyses = PhysChemAnalysis.query.filter(
            (PhysChemAnalysis.client.ilike(f'%{query}%')) |
            (PhysChemAnalysis.location.ilike(f'%{query}%')) |
            (PhysChemAnalysis.file_number.ilike(f'%{query}%'))
        ).order_by(PhysChemAnalysis.created_at.desc()).all()
        return jsonify([analysis.to_dict() for analysis in analyses]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# Microbiological Analysis Routes
@app.route('/api/micro', methods=['POST'])
def create_micro():
    try:
        data = request.json
        
        # Convert date strings to date objects
        date_collected = datetime.fromisoformat(data.get('dateCollected')).date() if data.get('dateCollected') else None
        date_analyzed = datetime.fromisoformat(data.get('dateAnalyzed')).date() if data.get('dateAnalyzed') else None
        date_submitted = datetime.fromisoformat(data.get('dateSubmitted')).date() if data.get('dateSubmitted') else None
        
        # Get signature flags - CORRECTED to read the right keys
        show_benjamin = data.get('showSignature', False)  # Read 'showSignature'
        show_eric = data.get('showEricSignature', False)  # Read 'showEricSignature'
        analyst_signature_scale = data.get('analystSignatureScale', 100)
        approver_signature_scale = data.get('approverSignatureScale', 100)
        app.logger.debug("Micro signature flags: show_benjamin=%s, show_eric=%s", show_benjamin, show_eric)
        
        analysis = MicrobiologicalAnalysis(
            client=data.get('client'),
            source=data.get('source'),
            location=data.get('location'),
            date_collected=date_collected,
            date_analyzed=date_analyzed,
            date_submitted=date_submitted,
            collected_by=data.get('collectedBy'),
            file_prefix=data.get('filePrefix'),
            file_number=data.get('fileNumber'),
            or_number=data.get('orNumber'),
            total_coliform=float(data.get('totalColiform')) if data.get('totalColiform') else None,
            e_coli=float(data.get('eColi')) if data.get('eColi') else None,
            fecal_coliform=float(data.get('fecalColiform')) if data.get('fecalColiform') else None,
            heterotrophic_plate_count=float(data.get('heterotrophicPlateCount')) if data.get('heterotrophicPlateCount') else None
        )
        
        db.session.add(analysis)
        db.session.commit()
        
        # Generate Excel file WITH signature flags
        excel_filename = create_micro_excel(
            analysis,
            show_benjamin,
            show_eric,
            analyst_signature_scale,
            approver_signature_scale
        )
        
        # Generate PDF from Excel
        pdf_filename = None
        if excel_filename:
            excel_path = os.path.join(app.config['UPLOAD_FOLDER'], 'micro', excel_filename)
            pdf_filename = convert_excel_to_pdf(excel_path)
        
        # Save filenames to database
        analysis.excel_file = excel_filename
        analysis.pdf_file = pdf_filename
        db.session.commit()
        
        return jsonify({
            'message': 'Microbiological analysis saved successfully!',
            'id': analysis.id,
            'excel_file': excel_filename,
            'pdf_file': pdf_filename
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/micro', methods=['GET'])
def get_all_micro():
    try:
        analyses = MicrobiologicalAnalysis.query.order_by(MicrobiologicalAnalysis.created_at.desc()).all()
        return jsonify([analysis.to_dict() for analysis in analyses]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/micro/<int:id>', methods=['GET'])
def get_micro(id):
    try:
        analysis = MicrobiologicalAnalysis.query.get_or_404(id)
        return jsonify(analysis.to_dict()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@app.route('/api/micro/search', methods=['GET'])
def search_micro():
    try:
        query = request.args.get('q', '')
        analyses = MicrobiologicalAnalysis.query.filter(
            (MicrobiologicalAnalysis.client.ilike(f'%{query}%')) |
            (MicrobiologicalAnalysis.location.ilike(f'%{query}%'))
        ).order_by(MicrobiologicalAnalysis.created_at.desc()).all()
        return jsonify([analysis.to_dict() for analysis in analyses]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# File upload routes
@app.route('/api/physchem/upload', methods=['POST'])
def upload_physchem_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if file and allowed_excel_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_filename = f"{timestamp}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'physchem', new_filename)
            file.save(filepath)
            
            return jsonify({
                'message': 'File uploaded successfully!',
                'filename': new_filename,
                'original_filename': filename
            }), 201
        else:
            return jsonify({'error': 'Invalid file type. Only .xlsx and .xls allowed'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/micro/upload', methods=['POST'])
def upload_micro_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if file and allowed_excel_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_filename = f"{timestamp}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'micro', new_filename)
            file.save(filepath)
            
            return jsonify({
                'message': 'File uploaded successfully!',
                'filename': new_filename,
                'original_filename': filename
            }), 201
        else:
            return jsonify({'error': 'Invalid file type. Only .xlsx and .xls allowed'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# List uploaded files
@app.route('/api/physchem/files', methods=['GET'])
def list_physchem_files():
    try:
        folder = os.path.join(app.config['UPLOAD_FOLDER'], 'physchem')
        files = []
        if os.path.exists(folder):
            for filename in os.listdir(folder):
                if allowed_excel_file(filename):
                    filepath = os.path.join(folder, filename)
                    file_stat = os.stat(filepath)
                    files.append({
                        'filename': filename,
                        'size': file_stat.st_size,
                        'uploaded_at': datetime.fromtimestamp(file_stat.st_mtime).isoformat()
                    })
        files.sort(key=lambda x: x['uploaded_at'], reverse=True)
        return jsonify(files), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/micro/files', methods=['GET'])
def list_micro_files():
    try:
        folder = os.path.join(app.config['UPLOAD_FOLDER'], 'micro')
        files = []
        if os.path.exists(folder):
            for filename in os.listdir(folder):
                if allowed_excel_file(filename):
                    filepath = os.path.join(folder, filename)
                    file_stat = os.stat(filepath)
                    files.append({
                        'filename': filename,
                        'size': file_stat.st_size,
                        'uploaded_at': datetime.fromtimestamp(file_stat.st_mtime).isoformat()
                    })
        files.sort(key=lambda x: x['uploaded_at'], reverse=True)
        return jsonify(files), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# Download file
# Download file
@app.route('/api/physchem/download/<filename>', methods=['GET'])
def download_physchem_file(filename):
    try:
        # Use absolute path
        base_dir = os.path.dirname(os.path.abspath(__file__))
        folder = os.path.join(base_dir, '..', 'uploads', 'physchem')
        folder = os.path.abspath(folder)  # Convert to absolute path
        
        app.logger.debug("PhysChem download lookup folder=%s file=%s exists=%s", folder, filename, os.path.exists(os.path.join(folder, filename)))
        
        return send_from_directory(folder, filename, as_attachment=True)
    except Exception as e:
        app.logger.exception("PhysChem download failed for file=%s", filename)
        return jsonify({'error': str(e)}), 404

@app.route('/api/micro/download/<filename>', methods=['GET'])
def download_micro_file(filename):
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        folder = os.path.join(base_dir, '..', 'uploads', 'micro')
        folder = os.path.abspath(folder)
        
        return send_from_directory(folder, filename, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404
    
# ==================== WATER TREATMENT ROUTES ====================

@app.route('/api/water-treatment', methods=['POST'])
def create_water_treatment():
    try:
        data = request.json
        
        # Convert datetime string to datetime object
        reading_datetime = datetime.fromisoformat(data.get('readingDatetime')) if data.get('readingDatetime') else None
        
        reading = WaterTreatmentReading(
            reading_datetime=reading_datetime,
            dam_level=float(data.get('damLevel')) if data.get('damLevel') else None,
            raw_water_turbidity=float(data.get('rawWaterTurbidity')) if data.get('rawWaterTurbidity') else None,
            clarified_water_phase1=float(data.get('clarifiedWaterPhase1')) if data.get('clarifiedWaterPhase1') else None,
            clarified_water_phase2=float(data.get('clarifiedWaterPhase2')) if data.get('clarifiedWaterPhase2') else None,
            filtered_water_phase1=float(data.get('filteredWaterPhase1')) if data.get('filteredWaterPhase1') else None,
            filtered_water_phase2=float(data.get('filteredWaterPhase2')) if data.get('filteredWaterPhase2') else None,
            pac_dosage=float(data.get('pacDosage')) if data.get('pacDosage') else None,
            alum_dosage=float(data.get('alumDosage')) if data.get('alumDosage') else None,
            notes=data.get('notes')
        )
        
        db.session.add(reading)
        db.session.commit()
        
        return jsonify({
            'message': 'Water treatment reading saved successfully!',
            'id': reading.id
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400


@app.route('/api/screen-data/live', methods=['GET'])
def get_live_screen_data():
    try:
        payload = _run_scrape_coroutine(_scrape_screen_data_live)
        return jsonify(payload), 200
    except Exception as e:
        app.logger.exception("Live screen data scrape failed")
        fallback_payload = _build_screen_data_fallback_payload(str(e))
        return jsonify(fallback_payload), 200


@app.route('/api/screen-data/history', methods=['GET'])
def get_screen_data_history():
    try:
        start_date_value = request.args.get('start_date')
        end_date_value = request.args.get('end_date')

        start_date = None
        end_date = None

        if start_date_value:
            try:
                start_date = datetime.strptime(start_date_value, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Invalid start_date format. Use YYYY-MM-DD.'}), 400

        if end_date_value:
            try:
                end_date = datetime.strptime(end_date_value, '%Y-%m-%d') + timedelta(days=1)
            except ValueError:
                return jsonify({'error': 'Invalid end_date format. Use YYYY-MM-DD.'}), 400

        grouped_history = _build_screen_data_history(start_date=start_date, end_date=end_date)
        return jsonify(grouped_history), 200
    except Exception as e:
        app.logger.exception('Failed to fetch screen data history')
        return jsonify({'error': str(e)}), 500


@app.route('/api/screen-data/history/missing-hours', methods=['GET'])
def get_screen_data_missing_hours():
    try:
        start_date_value = request.args.get('start_date')
        end_date_value = request.args.get('end_date')

        start_date = None
        end_date = None

        if start_date_value:
            try:
                start_date = datetime.strptime(start_date_value, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Invalid start_date format. Use YYYY-MM-DD.'}), 400

        if end_date_value:
            try:
                end_date = datetime.strptime(end_date_value, '%Y-%m-%d') + timedelta(days=1)
            except ValueError:
                return jsonify({'error': 'Invalid end_date format. Use YYYY-MM-DD.'}), 400

        payload = _build_missing_screen_data_hours(start_date=start_date, end_date=end_date)
        return jsonify(payload), 200
    except Exception as e:
        app.logger.exception('Failed to scan missing screen data hours')
        return jsonify({'error': str(e)}), 500


@app.route('/api/screen-data/history/manual-entries', methods=['POST'])
def save_screen_data_manual_entries():
    permission_error = _require_permission('edit_screen_data')
    if permission_error:
        return permission_error

    try:
        payload = request.get_json(silent=True) or {}
        entries = payload.get('entries')

        saved_count = _upsert_manual_screen_data_entries(entries)
        return jsonify({'message': f'Saved {saved_count} manual entr{("y" if saved_count == 1 else "ies")}.', 'savedCount': saved_count}), 200
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        app.logger.exception('Failed to save manual screen data entries')
        return jsonify({'error': str(e)}), 500


@app.route('/api/screen-data/last-chlorine-tank-change', methods=['GET', 'PUT'])
def manage_last_chlorine_tank_change():
    try:
        if request.method == 'GET':
            return jsonify({'date': _get_last_chlorine_tank_change()}), 200

        permission_error = _require_permission('edit_screen_data')
        if permission_error:
            return permission_error

        payload = request.get_json(silent=True) or {}
        date_value = (payload.get('date') or '').strip()

        if date_value and not re.fullmatch(r'\d{4}-\d{2}-\d{2}', date_value):
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

        saved_value = _set_last_chlorine_tank_change(date_value)
        return jsonify({'message': 'Last chlorine tank change updated.', 'date': saved_value}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/screen-data/last-active-dosing', methods=['GET', 'PUT'])
def manage_last_active_dosing():
    try:
        if request.method == 'GET':
            return jsonify({'value': _get_last_active_dosing()}), 200

        permission_error = _require_permission('edit_screen_data')
        if permission_error:
            return permission_error

        payload = request.get_json(silent=True) or {}
        value = (payload.get('value') or '').strip()

        if len(value) > 120:
            return jsonify({'error': 'Value is too long (max 120 characters).'}), 400

        saved_value = _set_last_active_dosing(value)
        return jsonify({'message': 'Last active dosing updated.', 'value': saved_value}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/water-treatment', methods=['GET'])
def get_all_water_treatment():
    try:
        readings = WaterTreatmentReading.query.order_by(WaterTreatmentReading.reading_datetime.desc()).all()
        return jsonify([reading.to_dict() for reading in readings]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/water-treatment/<int:id>', methods=['GET'])
def get_water_treatment(id):
    try:
        reading = WaterTreatmentReading.query.get_or_404(id)
        return jsonify(reading.to_dict()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@app.route('/api/water-treatment/search', methods=['GET'])
def search_water_treatment():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = WaterTreatmentReading.query
        
        if start_date:
            start_dt = datetime.fromisoformat(start_date)
            query = query.filter(WaterTreatmentReading.reading_datetime >= start_dt)
        
        if end_date:
            end_dt = datetime.fromisoformat(end_date)
            if len(end_date) == 10:
                end_dt = end_dt + timedelta(days=1)
                query = query.filter(WaterTreatmentReading.reading_datetime < end_dt)
            else:
                query = query.filter(WaterTreatmentReading.reading_datetime <= end_dt)
        
        readings = query.order_by(WaterTreatmentReading.reading_datetime.desc()).all()
        return jsonify([reading.to_dict() for reading in readings]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/water-treatment/advanced-search', methods=['GET'])
def advanced_search_water_treatment():
    try:
        # Get search parameters
        raw_turbidity_min = request.args.get('raw_turbidity_min', type=float)
        raw_turbidity_max = request.args.get('raw_turbidity_max', type=float)
        dam_level_min = request.args.get('dam_level_min', type=float)
        dam_level_max = request.args.get('dam_level_max', type=float)
        
        # Find matching raw water records
        query = WaterTreatmentReading.query
        
        if raw_turbidity_min is not None:
            query = query.filter(WaterTreatmentReading.raw_water_turbidity >= raw_turbidity_min)
        if raw_turbidity_max is not None:
            query = query.filter(WaterTreatmentReading.raw_water_turbidity <= raw_turbidity_max)
        if dam_level_min is not None:
            query = query.filter(WaterTreatmentReading.dam_level >= dam_level_min)
        if dam_level_max is not None:
            query = query.filter(WaterTreatmentReading.dam_level <= dam_level_max)
        
        raw_records = query.all()
        
        # For each raw water record, find the clarified water 2 hours later
        results = []
        for raw_record in raw_records:
            # Calculate time 2 hours later
            two_hours_later = raw_record.reading_datetime + timedelta(hours=2)
            
            # Find the closest record around that time (within 30 minutes)
            clarified_record = WaterTreatmentReading.query.filter(
                WaterTreatmentReading.reading_datetime >= two_hours_later - timedelta(minutes=30),
                WaterTreatmentReading.reading_datetime <= two_hours_later + timedelta(minutes=30)
            ).order_by(
                func.abs(
                    func.julianday(WaterTreatmentReading.reading_datetime) - 
                    func.julianday(two_hours_later)
                )
            ).first()
            
            # If we found a clarified record, calculate effectiveness
            if clarified_record and (clarified_record.clarified_water_phase1 or clarified_record.clarified_water_phase2):
                # Use the average of both phases or whichever is available
                clarified_avg = None
                if clarified_record.clarified_water_phase1 and clarified_record.clarified_water_phase2:
                    clarified_avg = (clarified_record.clarified_water_phase1 + clarified_record.clarified_water_phase2) / 2
                elif clarified_record.clarified_water_phase1:
                    clarified_avg = clarified_record.clarified_water_phase1
                elif clarified_record.clarified_water_phase2:
                    clarified_avg = clarified_record.clarified_water_phase2
                
                if clarified_avg is not None:
                    # Calculate reduction percentage
                    reduction_pct = 0
                    if raw_record.raw_water_turbidity and raw_record.raw_water_turbidity > 0:
                        reduction_pct = ((raw_record.raw_water_turbidity - clarified_avg) / raw_record.raw_water_turbidity) * 100
                    
                    results.append({
                        'raw_record': raw_record.to_dict(),
                        'clarified_record': clarified_record.to_dict(),
                        'clarified_avg': round(clarified_avg, 2),
                        'reduction_percentage': round(reduction_pct, 2),
                        'pac_dosage': raw_record.pac_dosage,
                        'alum_dosage': raw_record.alum_dosage,
                        'raw_turbidity': raw_record.raw_water_turbidity,
                        'dam_level': raw_record.dam_level,
                        'raw_datetime': raw_record.reading_datetime.isoformat(),
                        'clarified_datetime': clarified_record.reading_datetime.isoformat(),
                        'notes': raw_record.notes
                    })
        
        # Sort by lowest clarified water (most effective treatment)
        results.sort(key=lambda x: x['clarified_avg'])
        
        # Return top 10 most effective treatments
        top_10 = results[:10]
        app.logger.debug("Advanced search results: total=%s returned=%s", len(results), len(top_10))

        return jsonify({
            'total_found': len(results),
            'top_treatments': top_10
        }), 200
        
    except Exception as e:
        app.logger.exception("Advanced water treatment search failed")
        return jsonify({'error': str(e)}), 400

@app.route('/api/water-treatment/<int:id>', methods=['DELETE'])
def delete_water_treatment(id):
    try:
        reading = WaterTreatmentReading.query.get_or_404(id)
        db.session.delete(reading)
        db.session.commit()
        return jsonify({'message': 'Reading deleted successfully'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/water-treatment/download-daily/<date>', methods=['GET'])
def download_daily_report(date):
    try:
        # Parse date (format: YYYY-MM-DD)
        target_date = datetime.strptime(date, '%Y-%m-%d').date()
        
        # Get all records for that day
        records = WaterTreatmentReading.query.filter(
            func.date(WaterTreatmentReading.reading_datetime) == target_date
        ).order_by(WaterTreatmentReading.reading_datetime).all()
        
        if not records:
            return jsonify({'error': 'No records found for this date'}), 404
        
        # Generate Excel file
        date_info = target_date.strftime('%B %d, %Y')
        filename = create_water_treatment_excel_report(records, 'Daily', date_info)
        
        if filename:
            # Use absolute path
            directory = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], 'water_treatment'))
            file_path = os.path.join(directory, filename)
            
            app.logger.debug("Daily report file check path=%s exists=%s", file_path, os.path.exists(file_path))
            
            if not os.path.exists(file_path):
                return jsonify({'error': f'File not found at {file_path}'}), 500
            
            # Send the file
            return send_from_directory(
                directory,
                filename,
                as_attachment=True
            )
        else:
            return jsonify({'error': 'Failed to generate report'}), 500
            
    except Exception as e:
        app.logger.exception("Daily water treatment report download failed for date=%s", date)
        return jsonify({'error': str(e)}), 400

@app.route('/api/water-treatment/download-monthly/<year>/<month>', methods=['GET'])
def download_monthly_report(year, month):
    try:
        # Parse year and month
        year = int(year)
        month = int(month)
        
        # Get first and last day of month
        first_day = datetime(year, month, 1)
        last_day = datetime(year, month, monthrange(year, month)[1], 23, 59, 59)
        
        # Get all records for that month
        records = WaterTreatmentReading.query.filter(
            WaterTreatmentReading.reading_datetime >= first_day,
            WaterTreatmentReading.reading_datetime <= last_day
        ).order_by(WaterTreatmentReading.reading_datetime).all()
        
        if not records:
            return jsonify({'error': 'No records found for this month'}), 404
        
        # Generate Excel file
        date_info = first_day.strftime('%B %Y')
        filename = create_water_treatment_excel_report(records, 'Monthly', date_info)
        
        if filename:
            # Use absolute path
            directory = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], 'water_treatment'))
            file_path = os.path.join(directory, filename)
            
            app.logger.debug("Monthly report file check path=%s exists=%s", file_path, os.path.exists(file_path))
            
            if not os.path.exists(file_path):
                return jsonify({'error': f'File not found at {file_path}'}), 500
            
            # Send the file
            return send_from_directory(
                directory,
                filename,
                as_attachment=True
            )
        else:
            return jsonify({'error': 'Failed to generate report'}), 500
            
    except Exception as e:
        app.logger.exception("Monthly water treatment report download failed for year=%s month=%s", year, month)
        return jsonify({'error': str(e)}), 400

# ==================== LEAVE RECORDS ROUTES ====================

def create_cto_pdf(cto):
    """Generate PDF CTO slip"""
    try:
        filename = f"CTO_Slip_{cto.employee_name.replace(' ', '_')}_{cto.date_filed.strftime('%Y%m%d')}.pdf"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'leave_records', filename)
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Create PDF
        c = canvas.Canvas(filepath, pagesize=letter)
        width, height = letter
        
        # Header
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(width/2, height - 0.8*inch, "Republic of the Philippines")
        
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(width/2, height - 1.05*inch, "ZAMBOANGA CITY WATER DISTRICT")
        
        c.setFont("Helvetica", 10)
        c.drawCentredString(width/2, height - 1.25*inch, "Pilar St. Zamboanga City")
        
        # Title
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(width/2, height - 1.8*inch, "COMPENSATORY TIME - OFF (CTO)")
        
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(width/2, height - 2.05*inch, "APPLICATION SLIP")
        
        # Form fields
        y_position = height - 2.6*inch
        
        # Employee No and Date
        c.setFont("Helvetica", 10)
        c.drawString(1*inch, y_position, "EMPLOYEE NO.")
        c.setFont("Helvetica-Bold", 10)
        c.drawString(2.3*inch, y_position, str(cto.employee_no))
        c.line(2.3*inch, y_position - 2, 3.5*inch, y_position - 2)
        
        c.setFont("Helvetica", 10)
        c.drawString(4.5*inch, y_position, "Date :")
        c.setFont("Helvetica-Bold", 10)
        date_str = cto.date_filed.strftime('%d-%b-%y') if cto.date_filed else ''
        c.drawString(5.2*inch, y_position, date_str)
        c.line(5.2*inch, y_position - 2, 7*inch, y_position - 2)
        
        y_position -= 0.3*inch
        
        # Name
        c.setFont("Helvetica", 10)
        c.drawString(1*inch, y_position, "NAME :")
        c.setFont("Helvetica-Bold", 11)
        c.drawString(1.7*inch, y_position, cto.employee_name or '')
        c.line(1.7*inch, y_position - 2, 7*inch, y_position - 2)
        
        y_position -= 0.3*inch
        
        # Date Covered
        c.setFont("Helvetica", 10)
        c.drawString(1*inch, y_position, "DATE COVERED :")
        c.setFont("Helvetica-Bold", 10)
        c.drawString(2.2*inch, y_position, cto.date_covered_description or '')
        c.line(2.2*inch, y_position - 2, 7*inch, y_position - 2)
        
        y_position -= 0.3*inch
        
        # From and To
        c.setFont("Helvetica", 10)
        c.drawString(1*inch, y_position, "FROM")
        c.setFont("Helvetica-Bold", 10)
        from_str = cto.from_date.strftime('%d-%b') if cto.from_date else ''
        c.drawString(1.7*inch, y_position, from_str)
        c.line(1.7*inch, y_position - 2, 3.5*inch, y_position - 2)
        
        c.setFont("Helvetica", 10)
        c.drawString(4.5*inch, y_position, "TO :")
        c.setFont("Helvetica-Bold", 10)
        to_str = cto.to_date.strftime('%d-%b') if cto.to_date else ''
        c.drawString(5.2*inch, y_position, to_str)
        c.line(5.2*inch, y_position - 2, 7*inch, y_position - 2)
        
        y_position -= 0.3*inch
        
        # Total Hours
        c.setFont("Helvetica", 10)
        c.drawString(1*inch, y_position, "TOTAL NO. OF HOURS APPLIED")
        c.setFont("Helvetica-Bold", 10)
        hours_str = f"{int(cto.total_hours)} hours" if cto.total_hours else ''
        c.drawString(3.5*inch, y_position, hours_str)
        c.line(3.5*inch, y_position - 2, 7*inch, y_position - 2)
        
        y_position -= 0.6*inch
        
        # Signature of Applicant
        c.line(3*inch, y_position, 7*inch, y_position)
        c.setFont("Helvetica", 9)
        c.drawRightString(7*inch, y_position - 0.15*inch, "SIGNATURE OF APPLICANT")
        
        y_position -= 0.6*inch
        
        # Recommending Approval
        c.setFont("Helvetica", 10)
        c.drawString(1*inch, y_position, "RECOMMENDING APPROVAL:")
        
        y_position -= 0.5*inch
        
        # Signature line
        c.line(3*inch, y_position, 7*inch, y_position)
        
        y_position -= 0.25*inch
        
        # Name and Title
        c.setFont("Helvetica-Bold", 11)
        name = (cto.recommending_approval_name or '').upper()
        c.drawCentredString(5*inch, y_position, name)
        
        y_position -= 0.2*inch
        
        c.setFont("Helvetica", 10)
        title = cto.recommending_approval_title or ''
        c.drawCentredString(5*inch, y_position, title)
        
        c.save()
        app.logger.info("CTO PDF created: %s", filename)
        return filename
        
    except Exception as e:
        app.logger.exception("Error creating CTO PDF")
        return None

# ==================== EMPLOYEE ROUTES ====================

@app.route('/api/employees', methods=['GET'])
def get_employees():
    try:
        employees = Employee.query.order_by(Employee.employee_name).all()
        return jsonify([emp.to_dict() for emp in employees]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/employees/sync', methods=['POST'])
def sync_employees_from_analysis():
    try:
        # Manually add common employees
        common_names = ['Mark', 'Manuel Benjamin', 'Crispulo', 'Eric', 'Benjamin']
        
        employee_no = 1
        synced = 0
        
        for name in common_names:
            if not Employee.query.filter_by(employee_name=name).first():
                while Employee.query.filter_by(employee_no=str(employee_no).zfill(3)).first():
                    employee_no += 1
                
                emp = Employee(
                    employee_no=str(employee_no).zfill(3),
                    employee_name=name,
                    position='Laboratory Technician',
                    department='Water Quality Division'
                )
                db.session.add(emp)
                synced += 1
                employee_no += 1
        
        db.session.commit()
        return jsonify({'message': f'Synced {synced} employees!', 'total': Employee.query.count()}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/employees/update-numbers', methods=['POST'])
def update_employee_numbers():
    try:
        # Mapping of names to actual employee numbers
        employee_mapping = {
            'Manuel Benjamin Obsequio': '723',
            'Manuel Benjamin': '723',  # In case shorter name is stored
            'Allan Mark R. Ong': '724',
            'Allan Mark': '724',
            'Mark': '724',
            'Benjamin A. Lasola Jr.': '906',
            'Benjamin': '906',
            'Crispulo War Y. Indoc': '881',
            'Crispulo': '881',
            'Eric V. Salaritan': '096',
            'Eric': '096'
        }
        
        updated_count = 0
        
        for stored_name, emp_no in employee_mapping.items():
            employee = Employee.query.filter_by(employee_name=stored_name).first()
            if employee:
                employee.employee_no = emp_no
                updated_count += 1
                app.logger.info("Updated employee number: %s -> %s", stored_name, emp_no)
        
        db.session.commit()
        
        return jsonify({
            'message': f'Updated {updated_count} employee numbers!',
            'total_employees': Employee.query.count()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Failed to update employee numbers")
        return jsonify({'error': str(e)}), 400



# CTO Application Routes
@app.route('/api/cto-applications', methods=['POST'])
def create_cto_application():
    try:
        data = request.json
        
        cto = CTOApplication(
            employee_no=data.get('employeeNo'),
            employee_name=data.get('employeeName'),
            date_filed=datetime.strptime(data.get('dateFiled'), '%Y-%m-%d').date() if data.get('dateFiled') else None,
            date_covered_description=data.get('dateCoveredDescription'),
            from_date=datetime.strptime(data.get('fromDate'), '%Y-%m-%d').date() if data.get('fromDate') else None,
            to_date=datetime.strptime(data.get('toDate'), '%Y-%m-%d').date() if data.get('toDate') else None,
            total_hours=float(data.get('totalHours')) if data.get('totalHours') else 0,
            applicant_signature=data.get('applicantSignature'),
            recommending_approval_name=data.get('recommendingApprovalName'),
            recommending_approval_title=data.get('recommendingApprovalTitle'),
            recommending_signature=data.get('recommendingSignature')
        )
        
        db.session.add(cto)
        db.session.commit()
        
        filename = create_cto_pdf(cto)
        if filename:
            cto.excel_file = filename  # Using same field for PDF
            db.session.commit()

        return jsonify({
            'message': 'CTO application saved successfully!',
            'id': cto.id
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/cto-applications/download/<int:id>', methods=['GET'])
def download_cto_slip(id):
    try:
        cto = CTOApplication.query.get_or_404(id)
        
        if not cto.excel_file:
            # Generate if not exists
            filename = create_cto_pdf(cto)
            if filename:
                cto.excel_file = filename
                db.session.commit()
        
        if cto.excel_file:
            directory = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], 'leave_records'))
            return send_from_directory(directory, cto.excel_file, as_attachment=True)
        else:
            return jsonify({'error': 'Failed to generate CTO slip'}), 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 400

@app.route('/api/cto-applications', methods=['GET'])
def get_cto_applications():
    try:
        applications = CTOApplication.query.order_by(CTOApplication.date_filed.desc()).all()
        return jsonify([app.to_dict() for app in applications]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/cto-applications/<int:id>', methods=['DELETE'])
def delete_cto_application(id):
    try:
        cto = CTOApplication.query.get_or_404(id)
        
        # Delete associated files
        if cto.excel_file:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'leave_records', cto.excel_file)
            if os.path.exists(file_path):
                os.remove(file_path)
        
        db.session.delete(cto)
        db.session.commit()
        return jsonify({'message': 'CTO application deleted successfully'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# Leave Application Routes
@app.route('/api/leave-applications', methods=['POST'])
def create_leave_application():
    try:
        import json
        data = request.json
        
        leave = LeaveApplication(
            employee_name=data.get('employeeName'),
            date_filed=datetime.strptime(data.get('dateFiled'), '%Y-%m-%d').date() if data.get('dateFiled') else None,
            leave_types=json.dumps(data.get('leaveTypes', [])),
            other_leave_type=data.get('otherLeaveType'),
            from_date=datetime.strptime(data.get('fromDate'), '%Y-%m-%d').date() if data.get('fromDate') else None,
            to_date=datetime.strptime(data.get('toDate'), '%Y-%m-%d').date() if data.get('toDate') else None,
            day_off=data.get('dayOff'),
            applicant_signature=data.get('applicantSignature'),
            recommending_approval_name=data.get('recommendingApprovalName'),
            recommending_approval_title=data.get('recommendingApprovalTitle'),
            recommending_signature=data.get('recommendingSignature'),
            date_signed=datetime.strptime(data.get('dateSigned'), '%Y-%m-%d').date() if data.get('dateSigned') else None
        )
        
        db.session.add(leave)
        db.session.commit()
        
        return jsonify({
            'message': 'Leave application saved successfully!',
            'id': leave.id
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/leave-applications', methods=['GET'])
def get_leave_applications():
    try:
        applications = LeaveApplication.query.order_by(LeaveApplication.date_filed.desc()).all()
        return jsonify([app.to_dict() for app in applications]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/leave-applications/<int:id>', methods=['DELETE'])
def delete_leave_application(id):
    try:
        leave = LeaveApplication.query.get_or_404(id)
        
        # Delete associated files
        if leave.excel_file:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'leave_records', leave.excel_file)
            if os.path.exists(file_path):
                os.remove(file_path)
        
        db.session.delete(leave)
        db.session.commit()
        return jsonify({'message': 'Leave application deleted successfully'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# Leave Credits Routes
@app.route('/api/leave-credits', methods=['GET'])
def get_leave_credits():
    try:
        credits = LeaveCredits.query.all()
        return jsonify([credit.to_dict() for credit in credits]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/leave-credits/<employee_no>', methods=['GET'])
def get_employee_leave_credits(employee_no):
    try:
        credit = LeaveCredits.query.filter_by(employee_no=employee_no).first()
        if credit:
            return jsonify(credit.to_dict()), 200
        else:
            return jsonify({'error': 'Employee not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/leave-credits', methods=['POST'])
def create_or_update_leave_credits():
    try:
        data = request.json
        employee_no = data.get('employeeNo')
        
        # Check if employee already exists
        credit = LeaveCredits.query.filter_by(employee_no=employee_no).first()
        
        if credit:
            # Update existing
            credit.employee_name = data.get('employeeName', credit.employee_name)
            credit.vacation_leave = float(data.get('vacationLeave', credit.vacation_leave))
            credit.sick_leave = float(data.get('sickLeave', credit.sick_leave))
            credit.cto_hours = float(data.get('ctoHours', credit.cto_hours))
            message = 'Leave credits updated successfully!'
        else:
            # Create new
            credit = LeaveCredits(
                employee_no=employee_no,
                employee_name=data.get('employeeName'),
                vacation_leave=float(data.get('vacationLeave', 0)),
                sick_leave=float(data.get('sickLeave', 0)),
                cto_hours=float(data.get('ctoHours', 0))
            )
            db.session.add(credit)
            message = 'Leave credits created successfully!'
        
        db.session.commit()
        return jsonify({
            'message': message,
            'data': credit.to_dict()
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

if __name__ == '__main__':
    app_port = int(os.getenv('APP_PORT', '5100'))
    app.run(debug=True, host='0.0.0.0', port=app_port)
